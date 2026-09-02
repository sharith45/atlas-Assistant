import os
import sys
import json
import argparse
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule
from openpyxl.worksheet.datavalidation import DataValidation

# Configurar encoding en Windows para evitar UnicodeEncodeError en la consola
if sys.platform == "win32":
    try:
        sys.stdout.reconfigure(encoding="utf-8")
        sys.stderr.reconfigure(encoding="utf-8")
    except Exception:
        pass

def hex_to_clean(hex_str):
    """Limpia cadenas de color hex (quita # y asegura 6 caracteres)."""
    if not hex_str:
        return "1B365D"
    clean = hex_str.lstrip('#')
    if len(clean) == 3:
        clean = ''.join([c*2 for c in clean])
    return clean.upper()


def compile_json_to_xlsx(json_input, output_path):
    """
    Compila una especificación JSON (o ruta a archivo .json) en un libro de Excel (.xlsx)
    profesional con formato corporativo, soporte para grandes volúmenes de datos, KPIs,
    fórmulas, formato condicional, tablas y gráficos.

    Nota: La validación y reparación de JSON se realiza en orchestrator.compile_json_to_xlsx
    antes de llamar a esta función. Si se llama directamente, json_input puede ser la ruta
    al archivo .json o un dict ya parseado.
    """
    if isinstance(json_input, str):
        if os.path.exists(json_input):
            with open(json_input, "r", encoding="utf-8") as f:
                data = json.load(f)
        else:
            data = json.loads(json_input)
    else:
        data = json_input

    wb = openpyxl.Workbook()
    if wb.active:
        wb.remove(wb.active)

    settings = data.get("settings", {})
    theme_color = hex_to_clean(settings.get("theme_color", "1B365D"))
    zebra_color = hex_to_clean(settings.get("zebra_color", "F2F5F8"))
    font_name = settings.get("font_name", "Calibri")
    
    font_title = Font(name=font_name, size=16, bold=True, color=theme_color)
    font_subtitle = Font(name=font_name, size=10, italic=True, color="555555")
    font_header = Font(name=font_name, size=11, bold=True, color="FFFFFF")
    font_bold = Font(name=font_name, size=11, bold=True)
    font_norm = Font(name=font_name, size=11)
    font_kpi_label = Font(name=font_name, size=9, bold=True, color="555555")
    font_kpi_val = Font(name=font_name, size=20, bold=True, color=theme_color)

    fill_header = PatternFill(start_color=theme_color, end_color=theme_color, fill_type="solid")
    fill_zebra = PatternFill(start_color=zebra_color, end_color=zebra_color, fill_type="solid")
    fill_kpi = PatternFill(start_color="EBF3FA", end_color="EBF3FA", fill_type="solid")

    thin_side = Side(style='thin', color='CCCCCC')
    thin_border = Border(top=thin_side, bottom=thin_side, left=thin_side, right=thin_side)
    double_bottom_border = Border(top=thin_side, bottom=Side(style='double', color=theme_color))

    sheets_data = data.get("sheets", [])
    if not sheets_data:
        # Fallback si se pasa un formato simplificado de lista directa
        sheets_data = [{"name": "Datos", "tables": [{"headers": data.get("headers", []), "rows": data.get("rows", [])}]}]

    for s_idx, sheet_spec in enumerate(sheets_data):
        sheet_name = sheet_spec.get("name", f"Hoja {s_idx+1}")
        ws = wb.create_sheet(title=sheet_name[:31]) # Límite Excel 31 caracteres
        
        # Opciones de vista
        ws.views.sheetView[0].showGridLines = sheet_spec.get("show_gridlines", True)
        if "freeze_panes" in sheet_spec:
            ws.freeze_panes = sheet_spec["freeze_panes"]
            
        current_row = 1

        # Título y Subtítulo de Hoja
        if "title" in sheet_spec:
            ws.cell(row=current_row, column=1, value=sheet_spec["title"]).font = font_title
            current_row += 1
        if "subtitle" in sheet_spec:
            ws.cell(row=current_row, column=1, value=sheet_spec["subtitle"]).font = font_subtitle
            current_row += 1
        if "title" in sheet_spec or "subtitle" in sheet_spec:
            current_row += 1 # Espacio

        # Tarjetas KPI (si existen)
        kpis = sheet_spec.get("kpis", [])
        if kpis:
            kpi_row = current_row
            col_start = 1
            for kpi in kpis:
                c_lbl = ws.cell(row=kpi_row, column=col_start, value=kpi.get("label", "KPI"))
                c_lbl.font = font_kpi_label
                c_lbl.alignment = Alignment(horizontal="center", vertical="center")
                
                val_row = kpi_row + 1
                c_val = ws.cell(row=val_row, column=col_start, value=kpi.get("formula") or kpi.get("value", 0))
                c_val.font = font_kpi_val
                c_val.alignment = Alignment(horizontal="center", vertical="center")
                if "number_format" in kpi:
                    c_val.number_format = kpi["number_format"]

                # Unir celdas de tarjeta (2x2)
                col_end = col_start + kpi.get("width", 2) - 1
                ws.merge_cells(start_row=kpi_row, start_column=col_start, end_row=kpi_row, end_column=col_end)
                ws.merge_cells(start_row=val_row, start_column=col_start, end_row=val_row+1, end_column=col_end)
                
                # Aplicar relleno y bordes al KPI
                for r in range(kpi_row, val_row + 2):
                    for c in range(col_start, col_end + 1):
                        ws.cell(row=r, column=c).fill = fill_kpi
                        ws.cell(row=r, column=c).border = thin_border
                        
                col_start = col_end + 2 # Espacio entre tarjetas
            current_row += 4

        # Estructura de Celdas Directas (Grid)
        for cell_spec in sheet_spec.get("grid", []):
            ref = cell_spec.get("cell")
            if ref:
                c = ws[ref]
                c.value = cell_spec.get("value")
                if cell_spec.get("bold"):
                    c.font = font_bold
                if cell_spec.get("number_format"):
                    c.number_format = cell_spec["number_format"]

        # Tablas de Datos
        for tbl_spec in sheet_spec.get("tables", []):
            start_row = tbl_spec.get("start_row", current_row)
            start_col = tbl_spec.get("start_col", 1)
            
            headers = tbl_spec.get("headers", [])
            rows = tbl_spec.get("rows", [])
            col_formats = tbl_spec.get("column_formats", {})
            
            # Encabezados
            if headers:
                for c_off, h_text in enumerate(headers):
                    cell = ws.cell(row=start_row, column=start_col + c_off, value=h_text)
                    cell.font = font_header
                    cell.fill = fill_header
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    cell.border = thin_border
                start_row += 1

            # Filas de Datos
            r_idx = start_row
            for r_data in rows:
                is_even = (r_idx % 2 == 0)
                for c_off, val in enumerate(r_data):
                    col_num = start_col + c_off
                    cell = ws.cell(row=r_idx, column=col_num, value=val)
                    cell.font = font_norm
                    cell.border = thin_border
                    if is_even:
                        cell.fill = fill_zebra
                        
                    # Aplicar formato numérico si existe para la columna (1-indexed str o int)
                    fmt_key = str(c_off + 1)
                    if fmt_key in col_formats:
                        cell.number_format = col_formats[fmt_key]
                        cell.alignment = Alignment(horizontal="right")
                    elif isinstance(val, (int, float)):
                        cell.alignment = Alignment(horizontal="right")
                    elif isinstance(val, str) and val.startswith("="):
                        cell.alignment = Alignment(horizontal="right")
                    else:
                        cell.alignment = Alignment(horizontal="left")
                r_idx += 1

            # Fila de Totales
            if tbl_spec.get("total_row", False):
                ws.cell(row=r_idx, column=start_col, value="TOTAL").font = font_bold
                ws.cell(row=r_idx, column=start_col).border = double_bottom_border
                for c_off in range(1, len(headers)):
                    col_num = start_col + c_off
                    col_let = get_column_letter(col_num)
                    fmt_key = str(c_off + 1)
                    
                    # Generar fórmula de SUMA automática si es columna numérica
                    sum_formula = f"=SUM({col_let}{start_row}:{col_let}{r_idx-1})"
                    cell = ws.cell(row=r_idx, column=col_num, value=sum_formula)
                    cell.font = font_bold
                    cell.border = double_bottom_border
                    cell.alignment = Alignment(horizontal="right")
                    if fmt_key in col_formats:
                        cell.number_format = col_formats[fmt_key]
                r_idx += 1

            current_row = r_idx + 2

        # Gráficos
        for chart_spec in sheet_spec.get("charts", []):
            c_type = chart_spec.get("type", "column")
            chart = BarChart() if c_type in ["column", "bar"] else (LineChart() if c_type == "line" else PieChart())
            chart.title = chart_spec.get("title", "Gráfico")
            
            data_ref_str = chart_spec.get("data_range")
            cats_ref_str = chart_spec.get("cats_range")
            
            if data_ref_str:
                # Ejemplo data_range: "C8:C20"
                min_c, min_r, max_c, max_r = openpyxl.utils.range_boundaries(data_ref_str)
                data_ref = Reference(ws, min_col=min_c, min_row=min_r, max_col=max_c, max_row=max_r)
                chart.add_data(data_ref, titles_from_data=chart_spec.get("titles_from_data", True))
                
            if cats_ref_str:
                min_c, min_r, max_c, max_r = openpyxl.utils.range_boundaries(cats_ref_str)
                cats_ref = Reference(ws, min_col=min_c, min_row=min_r, max_col=max_c, max_row=max_r)
                chart.set_categories(cats_ref)
                
            chart.width = chart_spec.get("width", 15)
            chart.height = chart_spec.get("height", 9)
            ws.add_chart(chart, chart_spec.get("cell", "I8"))

        # Auto-ajustar ancho de columnas según contenido
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.value and not str(cell.value).startswith("="):
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    parent_dir = os.path.dirname(os.path.abspath(output_path))
    if parent_dir:
        os.makedirs(parent_dir, exist_ok=True)

    wb.save(output_path)
    print(f"✅ Documento Excel compilado exitosamente en: '{output_path}'")
    return output_path

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Compilador JSON a Excel (.xlsx) para ATLAS")
    parser.add_argument("json_path", help="Ruta al archivo JSON de especificación")
    parser.add_argument("output_path", help="Ruta de salida para el archivo .xlsx final")
    args = parser.parse_args()

    try:
        compile_json_to_xlsx(args.json_path, args.output_path)
    except Exception as e:
        print(f"❌ Error compilando archivo Excel: {e}", file=sys.stderr)
        sys.exit(1)
